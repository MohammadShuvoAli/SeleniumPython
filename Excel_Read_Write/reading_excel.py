import openpyxl

# File -> Workbook -> Sheets -> Rows -> Cells
path = "C:\\Users\\Shuvo\\Desktop\\testdata.xlsx"
workbook = openpyxl.load_workbook(path)
sheet = workbook["Sheet1"]

rows = sheet.max_row # count number of row in excel sheet
columns = sheet.max_column # count number of columns in excel sheet

# reading all the rows & and columns from excel sheet
for r in range(1, rows+1):
    for c in range(1, columns+1):
        print(sheet.cell(r, c).value, end="     ")
    print()
